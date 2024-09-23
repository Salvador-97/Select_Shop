import openpyxl
from tkinter import ttk

def menuExcel(pestaña):
    botonLimpiar = ttk.Button(
        pestaña,
        text='Limpiar'
    )
    botonLimpiar.pack()

# def limpiarExcel():
#     archivoEditable(bookPath)
#     sheet = book['entarimadoLote']
#     if(sheet.max_row > 1):
#         sheet.delete_rows()
#     cerrarExcel()

def archivoEditable(ruta):
    global bookPath
    bookPath = ruta
    global book 
    book = openpyxl.load_workbook(bookPath)
    # print(f'Funcion editExcel path: {bookPath}')

def ruta():
    return bookPath

def guardarExcel():
    book.save(bookPath)
    
def cerrarExcel():
    book.close()

def editarExcel(listaDatos):    
    sheet = book['entarimadoLote']
    print(f"Celda {sheet.max_row}")

    nuevaCelda = sheet.max_row + 1
    sheet.cell(row=nuevaCelda, column=1, value=listaDatos[0])
    sheet.cell(row=nuevaCelda, column=2, value=listaDatos[11])
    sheet.cell(row=nuevaCelda, column=3, value=listaDatos[1])

bookPath = ''
book = None